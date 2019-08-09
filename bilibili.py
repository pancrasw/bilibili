import requests
import time
import json
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import Workbook
import os
import datetime

class Bilibili:
    def __init__(self,cookie=None):
        if cookie!=None:
            self.cookie=cookie#单数保留字符串，复数保留字典
            self.cookies=json.loads('{\"'+cookie.replace('=','":"').replace('; ','","')+"\"}")
            self.__lastcall__1=0.0
        else:
            pass
    
    def bwait(self):
        while (time.time()-self.__lastcall__1<0.4):#每分钟最多150条，人为设置延迟
            time.sleep(0.1)#如果响应时间过长，导致间隔已超过0.4，则无需再进入此循环
    
    def find_upvideos(self,upid):#只需输入up主的uid的数值形式，返回为字典类型
        url='https://space.bilibili.com/ajax/member/getSubmitVideos?'
        count=-30#初始化为该数值是为了第一次能进去
        videos=[]
        para={
            'mid':str(upid),
            'pagesize':30,#每页浏览30个,20和40都不行，不知道为什么
            'page':0
        }
        while (len(videos)-count)==para['pagesize']:#即多读取的视频是不是30个
            count=len(videos)
            para['page']=para['page']+1
            self.bwait()
            r=requests.get(url,params=para)
            data=json.loads(r.text)
            for i in data['data']['vlist']:
                videos.append(i)
        return videos
    
    def throwlove(self,avid):#参数列表为vid数值类型
        a={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134",
            'Cookie': self.cookie,
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            'Origin': 'https://www.bilibili.com',
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        requestcontent={
            "aid": str(avid),
            'like':'1',
            'csrf': '4242f02d3a0c747b1a81ae161afe818a'
        }
        r=requests.post("https://api.bilibili.com/x/web-interface/archive/like",data=requestcontent,headers=a)
        try:
            r.raise_for_status()
            #print(r.text)#有需要可将该注释取消，查看情况
        except:
            print("网络错误")

    def pourlove(self,upid):#将某一up主的视频全部点赞
        videos=self.find_upvideos(upid)
        for i in videos:
            self.bwait()
            self.throwlove(int(i['aid']))
    
    def addcoin(self,avid):
        header={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36 Edge/17.17134",
            'Cookie': self.cookie,
            "Host": "api.bilibili.com",
            "Cache-Control": "no-cache",
            "Proxy-Connection": "keep-alive",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Origin": "https://www.bilibili.com",
            "Referer": "https://www.bilibili.com/video/av" + str(avid),
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9"
        }
        requestcontent={
            "aid": str(avid),
            "cross_domain": "true",
            "csrf": self.cookies['bili_jct'],
            "multiply": '1',
            "select_like": '1'
        }
        r=requests.post("https://api.bilibili.com/x/web-interface/coin/add",data=requestcontent,headers=header)
        try:
            r.raise_for_status()
            print(r.text)
        except:
            print("网络错误")               
            
    class video:
        def __init__(self, avid):
            self.vid=avid
            header={"Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Connection": "keep-alive",
            "Host": "www.bilibili.com",
            "Upgrade-Insecure-Requests": "1",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36"
            }
            self.r=requests.get("https://www.bilibili.com/video/av"+str(avid),headers=header)
            self.bs=BeautifulSoup(self.r.text,"html.parser")

        def gettag(self):
            adata=self.bs.find_all(class_="tag")
            str_list=[]
            for i in adata:
                bdata=i.find(target="_blank")
                str_list.append(bdata.string)
            return str_list
    
    class homepage:
        def __init__(self):
            self.r=requests.get("https://www.bilibili.com/")
            self.bs=BeautifulSoup(self.r.text,"html.parser")
        
        def getrecommend(self):
            adata=self.bs.find_all(class_="groom-module")
            rec_list=[]
            for i in adata:
                a={}
                bdata=i.find(target="_blank")        
                a["avid"]=bdata["href"][9:-1]
                a["title"]=bdata["title"]
                cdata=i.find(class_="author")
                a["author"]=cdata.string[4:len(cdata.string)]
                rec_list.append(a)
                a["uid"]=Bilibili().getuid(a["author"])
                ddata=i.find(class_="play")
                a["play"]=ddata.string[3:len(ddata.string)-1]
                a["time"]=datetime.datetime.now().strftime("%Y-%m-%d %X")
                a["tag"]=Bilibili().video(a["avid"]).gettag()
            return rec_list


    def getuid(self,name):
        url="https://search.bilibili.com/upuser?keyword="+name
        r=requests.get(url)
        bs=BeautifulSoup(r.text,"html.parser")
        adata=bs.find(class_="headline")
        bdata=adata.find(target="_blank")
        result=re.search(re.compile(r"[0-9]+"),bdata["href"])
        return result[0]
        
def excel_linker_for_recommend(list):
    try:
        wb = openpyxl.load_workbook(os.path.dirname(__file__).replace("\\",'/')+'/recommend for bilibili.xlsx')#此处写具体文档地址
        sheet=wb["Sheet"]
    except:
        wb=Workbook()
        sheet=wb.active
        sheet["A1"]="avid"
        sheet["B1"]="title"
        sheet["C1"]="author"
        sheet["D1"]="uid"
        sheet["E1"]="play(w)"
        sheet["F1"]="time"
        sheet["G1"]="tag"        
    for i in list:
        sheet["A"+str(sheet.max_row+1)]=i["avid"]
        sheet["B"+str(sheet.max_row)]=i["title"]
        sheet["C"+str(sheet.max_row)]=i["author"]
        sheet["D"+str(sheet.max_row)]=i["uid"]
        sheet["E"+str(sheet.max_row)]=i["play"]
        sheet["F"+str(sheet.max_row)]=i["time"]
        sheet["G"+str(sheet.max_row)]=str(i["tag"])
    wb.save(os.path.dirname(__file__).replace("\\",'/')+'/recommend for bilibili.xlsx')        
    
ho=Bilibili().homepage()    
excel_linker_for_recommend(ho.getrecommend())